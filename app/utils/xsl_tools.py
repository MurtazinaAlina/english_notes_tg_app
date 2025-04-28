"""
Функции для работы с xsl-файлами.
"""
from io import BytesIO
from typing import BinaryIO

import openpyxl
from openpyxl.styles import Alignment, Font
from sqlalchemy.ext.asyncio import AsyncSession

from app.utils.custom_bot_class import Bot
from app.database.db import DataBase
from app.settings import (
    INDEX_MIN_ROW, EXAMPLES_SEPARATOR, EXCEL_COLUMNS_STAT_SHEET, EXCEL_PERCENT_COLUMN_FORMATTING, EXCEL_ATTEMPTS,
    EXCEL_STATISTICS, EXCEL_COLUMNS_ATTEMPTS_SHEET, EXCEL_COLUMNS_VCB_SHEET, EXCEL_CONTEXT_COLOR, EXCEL_NOTES,
    EXCEL_COLUMNS_NOTES_SHEET, SYSTEM_SHEETS, PLUG_TEMPLATE, EXCEL_TABLE_OF_CONTENTS, TABLE_OF_CONTENTS_TITLE,
    FILENAME_STATISTICS, FILENAME_VOCABULARY, FILENAME_ALL_DATA
)


# Функция для создания пустого xsl-файла по указанному пути
async def _create_xls_file(path_to_xls_file: str) -> str:
    """
    Функция для создания пустого xsl-файла по указанному пути

    :param path_to_xls_file: Путь к xsl-файлу
    :return: Путь к xsl-файлу
    """
    # Создаём пустой xsl-файл и сохраняем его по указанному пути
    openpyxl.Workbook().save(path_to_xls_file)
    return path_to_xls_file


# Функция для импорта данных из xsl-файла
async def import_data_from_xls_file(
        session: AsyncSession, bot: Bot, chat_id: int, data_file: BinaryIO) -> int:
    """
    Функция для загрузки данных из xsl-файла с данными в базу.

    Строки с данными для парсинга начинаются с индекса=2 (выше хедер, устанавливается в файле settings).
    Системные листы - листы, данные из которых не должны быть импортированы в словарь (при наличии таковых).

    Структура файла:

    - Системные листы (не предполагается парсинг данных), названия зафиксированы в settings.py

    - Лист с заметками.
      Столбцы:
      A: id внутри листа | B: заголовок | C: текст заметки | D: примеры (separated by '\\n')

    - Листы с темами словаря. 1 лист = 1 тема.
      Столбцы:
      A: id внутри темы | B: word | C: transcription | D: translation | E: context examples (separated by '\\n')

    Примеры в context examples разделены '\\n'.

    При изменении структуры скорректируйте код!

    :param session: Пользовательская сессия
    :param bot: Объект бота
    :param chat_id: ID чата
    :param data_file: BinaryIO объект, содержащий xsl-файл с данными
    :return: Количество добавленных слов или None
    """
    added = 0                                                           # Количество добавленных/обновленных записей

    # Открываем файл и записываем его в переменную
    wb = openpyxl.load_workbook(BytesIO(data_file.read()))

    # Получаем названия всех листов
    all_sheets = wb.sheetnames

    # Получаем все темы пользователя из БД
    all_topics = await DataBase.get_all_topics(session, bot.auth_user_id.get(chat_id))

    # Итерируемся по всем листам
    for sheet in all_sheets:

        # Пропускаем системные листы, не обрабатываем данные из них
        if sheet in SYSTEM_SHEETS:
            continue

        # Импортируем данные заметок в таблицу Notes
        if sheet == EXCEL_NOTES:

            # Итерируемся по всем строкам листа с данными и извлекаем значения
            for row in wb[sheet].iter_rows(min_row=INDEX_MIN_ROW, values_only=True):
                title, text, examples = row[1], row[2], row[3]

                # Если строки закончились, выходим из обработки листа (Если title пустое, то строки закончились)
                if not title:
                    break

                # Форматируем данные
                title = title.strip()
                text = text.strip()

                # Проверяем, существует ли заметка в БД
                note_obj = await DataBase.get_note_by_data(session, title, text, bot.auth_user_id[chat_id])

                # Если заметка существует, разбиваем примеры по разделителю и проверяем на наличие в таблице Context
                if note_obj:
                    if examples:
                        new_examples = None
                        examples = [
                            example.strip() for example in examples.split(EXAMPLES_SEPARATOR) if example.strip()
                        ]
                        for example in examples:
                            context_obj = await DataBase.get_context_by_data(session, None, note_obj.id, example)

                            # Если примера нет в таблице Context, создаем его
                            if not context_obj:
                                await DataBase.create_context_example(
                                    session, {'context': example}, note_id=note_obj.id, word_id=None
                                )
                                new_examples = True

                        # Обновляем счётчик добавленных/изменённых заметок
                        if new_examples:
                            added += 1

                # Если заметки нет в БД, создаём ее
                else:
                    note_data = {'title': title, 'text': text, 'examples': PLUG_TEMPLATE}
                    new_note = await DataBase.create_new_note(session, bot.auth_user_id[chat_id], note_data)

                    # Обновляем счётчик добавленных/изменённых заметок
                    if new_note:
                        added += 1

                    # Если есть примеры, разбиваем их по разделителю и создаем записи в таблице Context
                    if examples:
                        examples = [
                            example.strip() for example in examples.split(EXAMPLES_SEPARATOR) if example.strip()
                        ]
                        for example in examples:
                            await DataBase.create_context_example(
                                session, {'context': example}, note_id=new_note.id, word_id=None
                            )

        # Импортируем данные словаря в таблицу WordPhrase
        else:

            # Если названия листа нет в таблице Topic, создаём его в БД
            if sheet not in [topic.name for topic in all_topics]:
                topic = await DataBase.create_topic(session, {'name': sheet}, bot.auth_user_id[chat_id])
            else:
                topic = [topic for topic in all_topics if topic.name == sheet][0]

            # Итерируемся по всем строкам листа с данными и извлекаем значения
            for row in wb[sheet].iter_rows(min_row=INDEX_MIN_ROW, values_only=True):
                word, transcription, translate, examples = row[1], row[2], row[3], row[4]

                # Если строки закончились, выходим из обработки листа (Если word пустое, то строки закончились)
                if not word:
                    break

                # Форматируем данные (Обрабатываем None + лишние символы)
                word = word.strip()
                transcription = transcription.strip() if transcription else ''
                translate = translate.strip() if translate else ''

                # Проверяем, существует ли такая запись WordPhrase в БД (допускается расширение перевода)
                word_phrase_obj = await DataBase.get_word_phrase_by_data(session, word, translate, topic.id)

                # Если заметка существует, проверяем идентичность перевода. Если перевод был расширен, перезаписываем
                if word_phrase_obj:
                    is_updated = False
                    if translate != word_phrase_obj.translate:
                        upd_data = {'translate': translate}
                        is_updated = await DataBase.update_word_phrase(session, word_phrase_obj.id, upd_data)

                        # Обновляем счётчик добавленных/изменённых заметок
                        if is_updated:
                            added += 1

                    # Если есть примеры, разбиваем их по разделителю и проверяем на наличие в таблице Context
                    if examples:
                        new_examples = None
                        examples = [
                            example.strip() for example in examples.split(EXAMPLES_SEPARATOR) if example.strip()
                        ]
                        for example in examples:
                            context_obj = await DataBase.get_context_by_data(session, word_phrase_obj.id, None, example)

                            # Если примера нет в таблице Context, создаем его
                            if not context_obj:
                                await DataBase.create_context_example(
                                    session, {'context': example}, note_id=None, word_id=word_phrase_obj.id
                                )
                                new_examples = True

                        # Обновляем счётчик добавленных/изменённых заметок (если уже не был обновлен перевод)
                        if new_examples and not is_updated:
                            added += 1

                # Если записи WordPhrase нет в БД, создаём ее
                else:
                    new_word = await DataBase.create_word_phrase(
                        session, {
                            'word': word, 'transcription': transcription, 'translate': translate, 'topic': topic.id
                        }
                    )

                    # Увеличиваем счетчик добавленных слов
                    if new_word:
                        added += 1

                        # Если есть примеры, разбиваем их по разделителю и создаем записи в таблице Context
                        if examples:
                            examples = [c.strip() for c in examples.split(EXAMPLES_SEPARATOR) if c.strip()]
                            for example in examples:
                                await DataBase.create_context_example(
                                    session, {'context': example}, note_id=None, word_id=new_word.id
                                )
    return added


# Экспорт данных словаря пользователя + заметок в xsl-файл
async def export_vcb_data_to_xls_file(
        session: AsyncSession, bot: Bot, chat_id: int, path_to_xls_file: str = FILENAME_VOCABULARY
) -> str:
    """
    Функция создаёт сводный xsl-файл и экспортирует в него все данные WordPhrase словаря пользователя + заметки Notes.

    :param session: Пользовательская сессия
    :param bot: Объект бота
    :param chat_id: ID чата
    :param path_to_xls_file: Путь к xsl-файлу
    :return: Относительный путь к сформированному xsl-файлу (в корне проекта по умолчанию)
    """

    # Создаём сводный xsl-файл для экспорта словаря
    await _create_xls_file(path_to_xls_file)

    # Загружаем его в переменную
    wb = openpyxl.load_workbook(path_to_xls_file)

    # Делаем лист для оглавления и форматируем его
    ws_cont = wb.active
    ws_cont.title = EXCEL_TABLE_OF_CONTENTS
    ws_cont.sheet_view.showGridLines = False
    ws_cont.column_dimensions['A'].width = 40
    ws_cont["A1"] = TABLE_OF_CONTENTS_TITLE
    ws_cont["A1"].font = Font(bold=True)
    row_cont = 2

    # Итерируемся по всем темам пользователя в БД
    all_topics = await DataBase.get_all_topics(session, bot.auth_user_id.get(chat_id))
    for topic in all_topics:

        # Забираем все слова этой темы из БД
        all_words = await DataBase.get_user_word_phrases(
            session, bot.auth_user_id.get(chat_id), topic.id, ordering_asc=True
        )

        # Создаём новый лист в xsl-файле с названием темы
        ws = wb.create_sheet(topic.name)

        # Добавляем лист в оглавление
        ws_cont[f"A{row_cont}"] = topic.name
        ws_cont[f"A{row_cont}"].font = Font(
            color=EXCEL_CONTEXT_COLOR,
            bold=True,
            italic=True,
            underline=None
        )
        ws_cont[f"A{row_cont}"].hyperlink = f"#'{topic.name}'!A1"
        row_cont += 1

        # Создаём хедер таблицы с данными словаря, форматируем, выравниваем, задаём ширину столбцов
        for col, value in EXCEL_COLUMNS_VCB_SHEET.items():
            ws[f'{col}1'] = value['header']
            ws[f'{col}1'].font = Font(bold=True)
            ws[f'{col}1'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws.column_dimensions[col].width = value['width']

        # Заполняем столбцы данными записей WordPhrase
        row_index = INDEX_MIN_ROW
        word_id = 1                                                           # ID слова внутри листа
        for word in all_words:
            ws.append([word_id, word.word, word.transcription, word.translate])
            word_id += 1

            # Добавляем примеры контекста в столбец Е
            cell = ws[f'E{row_index}']
            cell.value = ''
            for context in word.context:
                if cell.value == '':
                    cell.value = f'{context.example}'
                else:
                    cell.value += f'\n{context.example}'
            row_index += 1

        # Форматируем страницу

        # Устанавливаем перенос строк в ячейках, выравнивание, шрифт
        for column in ['A', 'B', 'C', 'D', 'E']:
            for row in range(INDEX_MIN_ROW, ws.max_row + 1):
                cell = ws[f"{column}{row}"]

                # Форматируем строки таблицы (кроме заголовков)
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

                # Форматируем столбец E (контекст)
                if column == 'E':
                    cell.font = Font(italic=True, color=EXCEL_CONTEXT_COLOR)

    # ЗАМЕТКИ

    # Забираем из БД все заметки пользователя
    all_notes = await DataBase.get_user_notes(session, bot.auth_user_id.get(chat_id), ordering_asc=True)

    # Создаём новый лист в xsl-файле для заметок и перемещаем его в начало
    ws = wb.create_sheet(EXCEL_NOTES)
    wb._sheets.insert(1, wb._sheets.pop(wb._sheets.index(ws)))              # Перемещаем лист в начало

    # Создаём хедер таблицы с данными заметок, форматируем, выравниваем, задаём ширину столбцов
    for col, value in EXCEL_COLUMNS_NOTES_SHEET.items():
        ws[f'{col}1'] = value['header']
        ws[f'{col}1'].font = Font(bold=True)
        ws[f'{col}1'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        ws.column_dimensions[col].width = value['width']

    # Заполняем столбцы данными записей Notes
    row_index = INDEX_MIN_ROW
    note_id = 1                                                           # ID слова внутри листа
    for note in all_notes:
        ws.append([note_id, note.title, note.text])
        note_id += 1

        # Добавляем примеры контекста в столбец D
        cell = ws[f'D{row_index}']
        cell.value = ''
        for context in note.examples:
            if cell.value == '':
                cell.value = f'{context.example}'
            else:
                cell.value += f'\n{context.example}'
        row_index += 1

    # Форматируем страницу

    # Устанавливаем перенос строк в ячейках, выравнивание, шрифт
    for column in ['A', 'B', 'C', 'D']:
        for row in range(INDEX_MIN_ROW, ws.max_row + 1):
            cell = ws[f"{column}{row}"]

            # Форматируем строки таблицы (кроме заголовков)
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

            # Форматируем столбец D (контекст)
            if column == 'D':
                cell.font = Font(italic=True, color=EXCEL_CONTEXT_COLOR)

    # Сохраняем изменения в xsl-файле
    wb.save(path_to_xls_file)
    return path_to_xls_file


# Функция добавления статистики пользователя в wb
async def _add_statistic_data_to_wb(session: AsyncSession, user_id: int, reports: list, path_to_xls_file: str,
                                    stat_only: bool = True) -> str:
    """
    Функция добавления статистики пользователя в wb.
    Используется для формирования отдельного отчёта статистики или добавления статистики в xls со всеми данными
    пользователя.

    :param session: Пользовательская сессия
    :param user_id: ID пользователя User
    :param reports: Список отчетов Report
    :param path_to_xls_file: Путь к xsl-файлу
    :param stat_only: Флаг, указывающий, нужно ли удалять лист 'Sheet'
    :return: Путь к xsl-файлу
    """

    # Загружаем рабочую книгу xsl в переменную
    wb = openpyxl.load_workbook(path_to_xls_file)

    # Создаём новый лист в xsl-файле для записи отчетов
    ws = wb.create_sheet(EXCEL_STATISTICS)

    # Создаём хедер таблицы с отчётами Report, форматируем, выравниваем, задаём ширину столбцов
    for col, value in EXCEL_COLUMNS_STAT_SHEET.items():
        ws[f'{col}1'] = value['header']
        ws[f'{col}1'].font = Font(bold=True)
        ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[col].width = value['width']

    # Заполняем таблицу отчетами Report
    row_index = INDEX_MIN_ROW
    user_report_counter = 1
    for report in reports:
        topic_name = report.topic_name if report.topic_name else '-'
        result_percentage = report.result_percentage / 100
        report_data = [
            user_report_counter, report.created, report.correct_attempts, report.total_attempts, result_percentage,
            topic_name, report.total_words, report.test_type
        ]
        for col, value in zip(EXCEL_COLUMNS_STAT_SHEET.keys(), report_data):
            ws[f'{col}{row_index}'] = value
        row_index += 1
        user_report_counter += 1

    # Форматируем отображение в столбце с процентами %
    for row in range(INDEX_MIN_ROW, 2000):
        ws[f'E{row}'].number_format = EXCEL_PERCENT_COLUMN_FORMATTING

    # Создаём лист с отчетом попыток прохождения тестов пользователем
    ws = wb.create_sheet(EXCEL_ATTEMPTS)

    # Создаём хедер таблицы с отчётами Attempt, форматируем, выравниваем, задаём ширину столбцов
    for col, value in EXCEL_COLUMNS_ATTEMPTS_SHEET.items():
        ws.column_dimensions[col].width = value['width']
        ws[f'{col}1'] = value['header']
        ws[f'{col}1'].font = Font(bold=True)
        ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')

    # Заполняем таблицу попытками Attempt
    attempts = await DataBase.get_user_attempts(session, user_id)
    row_index = INDEX_MIN_ROW
    user_attempt_counter = 1
    for attempt in attempts:
        attempt.result = 'Верно' if attempt.result == 'correct' else 'Неверно'
        attempt_data = [
            user_attempt_counter, attempt.created, attempt.word_text, attempt.result, attempt.report_id,
            attempt.test_type
        ]

        # Заполняем строку попытки
        for col, value in zip(EXCEL_COLUMNS_ATTEMPTS_SHEET.keys(), attempt_data):
            ws[f'{col}{row_index}'] = value

        user_attempt_counter += 1
        row_index += 1

    # Если это экспорт только статистики, удаляем ненужный системный лист
    if stat_only:
        try:
            wb.remove(wb['Sheet'])
        except (Exception,):
            pass

    # Сохраняем изменения в xsl-файле
    wb.save(path_to_xls_file)
    return path_to_xls_file


# Функция экспорта статистики пользователя в xsl-файл
async def export_statistic_data_to_xls(session: AsyncSession, user_id: int, reports: list) -> str:
    """
    Функция экспорта статистики пользователя в xsl-файл.

    :param session: Пользовательская сессия
    :param user_id: ID пользователя User.id
    :param reports: Список отчетов пользователя
    :return: Путь к сохранённому xsl-файлу
    """

    # Создаём сводный xsl-файл для экспорта статистики
    await _create_xls_file(FILENAME_STATISTICS)

    # Добавляем данные статистики в файл
    await _add_statistic_data_to_wb(session, user_id, reports, FILENAME_STATISTICS)
    return FILENAME_STATISTICS


# Функция экспорта всех данных пользователя в xsl-файл (словарь + заметки + статистика)
async def export_all_user_data_to_xls(session: AsyncSession, bot: Bot, chat_id: int, user_id: int, reports: list) \
        -> str:
    """
    Функция экспорта всех данных пользователя в xsl-файл (словарь + заметки + статистика).

    :param session: Пользовательская сессия
    :param bot: Объект бота
    :param chat_id: ID чата
    :param user_id: ID пользователя User.id
    :param reports: Список отчетов пользователя
    :return: Путь к сохранённому xsl-файлу
    """

    # Создаём сводный xsl-файл и заполняем его данными словаря и заметок
    await export_vcb_data_to_xls_file(session, bot, chat_id, FILENAME_ALL_DATA)

    # Добавляем данные статистики в файл
    await _add_statistic_data_to_wb(session, user_id, reports, FILENAME_ALL_DATA, stat_only=False)

    # Перемещаем листы со статистикой в начало, после листа оглавления (темы проще находить через ссылки в оглавлении)
    wb = openpyxl.load_workbook(FILENAME_ALL_DATA)
    wb._sheets.insert(1, wb._sheets.pop(wb._sheets.index(wb[EXCEL_STATISTICS])))
    wb._sheets.insert(1, wb._sheets.pop(wb._sheets.index(wb[EXCEL_ATTEMPTS])))
    wb.save(FILENAME_ALL_DATA)

    return FILENAME_ALL_DATA
